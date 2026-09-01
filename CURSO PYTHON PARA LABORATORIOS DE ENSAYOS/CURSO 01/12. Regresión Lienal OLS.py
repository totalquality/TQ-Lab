
# ======================================================================
# REGRESIÓN LINEAL SIMPLE POR MÍNIMOS CUADRADOS ORDINARIOS (OLS)
# Aplicación: validación de métodos químicos cuantitativos
#
#   1. Ecuación de regresión
#   2. Coeficientes: Coef., EE, T, p y FIV
#   3. Resumen del modelo: S, R², R² ajustado y R² predicho
#   4. ANOVA completo
#   5. Lack-of-Fit + Error puro cuando existen réplicas
#   6. Ajustes y diagnóstico de observaciones poco comunes
#   7. Supuestos OLS:
#        - Normalidad de residuales: Anderson-Darling
#        - Independencia: Residuales vs. orden + Durbin-Watson
#        - Homocedasticidad: Residuales vs. ajustados + Breusch-Pagan
#   8. Gráficas modernas y de alta resolución
#   9. Exportación de resultados a Excel
#
# IMPORTANTE:
# Un R² elevado o un ANOVA significativo NO demuestran por sí solos
# que el modelo lineal sea adecuado. En validación se deben revisar
# simultáneamente los criterios del modelo y sus residuales.
# ======================================================================


# ======================================================================
# 0. IMPORTACIONES
# ======================================================================

import warnings
warnings.filterwarnings("ignore")

from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator

try:
    from IPython.display import display
except ImportError:
    display = None

import statsmodels.api as sm

from scipy import stats
from statsmodels.stats.diagnostic import normal_ad, het_breuschpagan
from statsmodels.stats.stattools import durbin_watson


# ======================================================================
# 1. CONFIGURACIÓN
# ======================================================================

# El archivo se busca en la misma carpeta donde se ejecuta el notebook.
ARCHIVO = "linealidad.xlsx"

# Si se deja None, Python utilizará automáticamente la primera hoja.
HOJA = None

ALPHA = 0.05

# Nombre del archivo final de resultados
ARCHIVO_SALIDA = "resultado_OLS_linealidad.xlsx"


# ======================================================================
# 2. ESTILO VISUAL
# ======================================================================
#
# Buscamos una presentación más moderna que la salida tradicional
# de Minitab, pero manteniendo una apariencia técnica y profesional.
# ======================================================================

plt.rcParams.update({
    "figure.dpi": 120,
    "savefig.dpi": 220,
    "font.family": "DejaVu Sans",
    "axes.titlesize": 14,
    "axes.titleweight": "bold",
    "axes.labelsize": 11,
    "axes.edgecolor": "#263238",
    "axes.linewidth": 1.0,
    "xtick.labelsize": 9.5,
    "ytick.labelsize": 9.5,
    "legend.fontsize": 9.5,
    "figure.facecolor": "white",
    "axes.facecolor": "#FBFCFE",
})


# ======================================================================
# 3. FUNCIONES AUXILIARES
# ======================================================================

def fmt(x, d=4):
    """Formato numérico cómodo para mostrar resultados."""
    if x is None:
        return "NA"

    try:
        if not np.isfinite(x):
            return "NA"
    except TypeError:
        return "NA"

    return f"{x:.{d}f}"


def fmt_p(p):
    """Formato de p-valor."""
    if p is None or not np.isfinite(p):
        return "NA"

    if p < 0.000001:
        return "< 0.000001"

    if p < 0.001:
        return f"{p:.6f}"

    return f"{p:.4f}"


def cargar_datos(archivo, hoja=None):
    """
    Lee el Excel y detecta automáticamente las dos primeras
    columnas numéricas.

    Esto permite cambiar los nombres de las columnas sin tener
    que modificar el resto del programa.
    """

    archivo_path = Path(archivo)

    if not archivo_path.exists():
        raise FileNotFoundError(
            f"No se encontró '{archivo}'. "
            "Coloca el Excel en la misma carpeta del notebook."
        )

    xls = pd.ExcelFile(archivo)

    if hoja is None:
        hoja_real = xls.sheet_names[0]
    else:
        if hoja not in xls.sheet_names:
            raise ValueError(
                f"La hoja '{hoja}' no existe.\n"
                f"Hojas disponibles: {xls.sheet_names}"
            )
        hoja_real = hoja

    # Primero intentamos leer normalmente.
    raw = pd.read_excel(
        archivo,
        sheet_name=hoja_real
    )

    # Identificar columnas que contienen datos numéricos.
    columnas_numericas = []

    for columna in raw.columns:

        serie = pd.to_numeric(
            raw[columna],
            errors="coerce"
        )

        if serie.notna().sum() >= 3:
            columnas_numericas.append(columna)

    if len(columnas_numericas) < 2:

        raise ValueError(
            "No se encontraron al menos dos columnas numéricas.\n"
            "El Excel debe contener una variable X y una variable Y."
        )

    # Utilizamos las dos primeras columnas numéricas.
    x_columna = columnas_numericas[0]
    y_columna = columnas_numericas[1]

    datos = pd.DataFrame({
        "X": pd.to_numeric(
            raw[x_columna],
            errors="coerce"
        ),
        "Y": pd.to_numeric(
            raw[y_columna],
            errors="coerce"
        )
    })

    datos = datos.dropna(
        subset=["X", "Y"]
    ).reset_index(drop=True)

    datos["Obs"] = np.arange(
        1,
        len(datos) + 1
    )

    return datos, hoja_real, x_columna, y_columna


def calcular_lack_of_fit(datos, modelo):
    """
    Descompone el error:

        SSE = SS_LOF + SS_PE

    donde:

        SS_LOF = falta de ajuste
        SS_PE  = error puro

    El error puro se obtiene utilizando las réplicas existentes
    dentro de cada nivel de X.

    Para OLS simple:

        p = 2

    (intercepto y pendiente).
    """

    grupos = (
        datos
        .groupby("X")["Y"]
        .agg(
            n="size",
            media="mean"
        )
        .reset_index()
    )

    n = len(datos)
    g = len(grupos)
    p = 2

    # --------------------------------------------------------------
    # Error puro
    # --------------------------------------------------------------

    SS_PE = 0.0

    for _, grupo in datos.groupby("X"):

        media_grupo = grupo["Y"].mean()

        SS_PE += np.sum(
            (grupo["Y"] - media_grupo) ** 2
        )

    # --------------------------------------------------------------
    # Error total del modelo
    # --------------------------------------------------------------

    SSE = np.sum(
        modelo.resid ** 2
    )

    # Por seguridad numérica
    SS_LOF = max(
        0.0,
        SSE - SS_PE
    )

    df_PE = n - g
    df_LOF = g - p

    MS_PE = (
        SS_PE / df_PE
        if df_PE > 0
        else np.nan
    )

    MS_LOF = (
        SS_LOF / df_LOF
        if df_LOF > 0
        else np.nan
    )

    if MS_PE > 0:

        F_LOF = MS_LOF / MS_PE

        p_LOF = stats.f.sf(
            F_LOF,
            df_LOF,
            df_PE
        )

    else:

        F_LOF = np.nan
        p_LOF = np.nan

    # ¿Existe al menos un nivel con réplicas?
    hay_replicas = bool(
        (grupos["n"] > 1).any()
    )

    return {
        "n": n,
        "g": g,
        "p": p,
        "SSE": SSE,
        "SS_PE": SS_PE,
        "SS_LOF": SS_LOF,
        "df_PE": df_PE,
        "df_LOF": df_LOF,
        "MS_PE": MS_PE,
        "MS_LOF": MS_LOF,
        "F_LOF": F_LOF,
        "p_LOF": p_LOF,
        "hay_replicas": hay_replicas,
        "grupos": grupos
    }


def calcular_anova(modelo):
    """ANOVA de regresión simple."""

    n = int(modelo.nobs)

    df_reg = 1
    df_error = n - 2
    df_total = n - 1

    SS_reg = modelo.ess
    SS_error = modelo.ssr
    SS_total = modelo.centered_tss

    MS_reg = SS_reg / df_reg
    MS_error = SS_error / df_error

    F = MS_reg / MS_error

    p = stats.f.sf(
        F,
        df_reg,
        df_error
    )

    tabla = pd.DataFrame({

        "Fuente": [
            "Regresión",
            "X",
            "Error",
            "Total"
        ],

        "GL": [
            df_reg,
            1,
            df_error,
            df_total
        ],

        "SC": [
            SS_reg,
            SS_reg,
            SS_error,
            SS_total
        ],

        "MC": [
            MS_reg,
            MS_reg,
            MS_error,
            np.nan
        ],

        "F": [
            F,
            F,
            np.nan,
            np.nan
        ],

        "p": [
            p,
            p,
            np.nan,
            np.nan
        ]
    })

    return tabla


def obtener_diagnosticos(modelo, datos):
    """
    Calcula los principales indicadores de diagnóstico
    de observaciones poco comunes.
    """

    influencia = modelo.get_influence()

    datos_d = datos.copy()

    datos_d["Ajuste"] = modelo.fittedvalues

    # El error estándar del ajuste se obtiene del objeto
    # de predicción del modelo.
    pred_diag = modelo.get_prediction(
        sm.add_constant(datos["X"])
    )

    frame_diag = pred_diag.summary_frame(
        alpha=ALPHA
    )

    datos_d["EE ajuste"] = (
        frame_diag["mean_se"].values
    )

    datos_d["Resid"] = modelo.resid

    datos_d["Resid est"] = (
        influencia.resid_studentized_internal
    )

    datos_d["Resid est externo"] = (
        influencia.resid_studentized_external
    )

    datos_d["Leverage"] = (
        influencia.hat_matrix_diag
    )

    datos_d["Cook"] = (
        influencia.cooks_distance[0]
    )

    datos_d["DFFITS"] = (
        influencia.dffits[0]
    )

    # --------------------------------------------------------------
    # IC 95 % de la media
    # --------------------------------------------------------------

    pred = modelo.get_prediction(
        sm.add_constant(datos["X"])
    )

    frame = pred.summary_frame(
        alpha=ALPHA
    )

    datos_d["IC95 media inferior"] = (
        frame["mean_ci_lower"].values
    )

    datos_d["IC95 media superior"] = (
        frame["mean_ci_upper"].values
    )

    # --------------------------------------------------------------
    # Intervalo de predicción 95 %
    # --------------------------------------------------------------

    datos_d["IP95 inferior"] = (
        frame["obs_ci_lower"].values
    )

    datos_d["IP95 superior"] = (
        frame["obs_ci_upper"].values
    )

    n = len(datos)
    p = 2

    # Umbrales orientativos.
    # Se muestran como diagnóstico, no como "criterios universales".
    umbral_resid = 2
    umbral_leverage = 2 * p / n
    umbral_cook = 4 / n

    datos_d["Marca R"] = np.where(
        np.abs(datos_d["Resid est"]) > umbral_resid,
        "R",
        ""
    )

    datos_d["Marca leverage"] = np.where(
        datos_d["Leverage"] > umbral_leverage,
        "L",
        ""
    )

    datos_d["Marca Cook"] = np.where(
        datos_d["Cook"] > umbral_cook,
        "C",
        ""
    )

    return (
        datos_d,
        umbral_resid,
        umbral_leverage,
        umbral_cook
    )


# ======================================================================
# 4. CARGAR DATOS
# ======================================================================

datos, hoja_real, nombre_x, nombre_y = cargar_datos(
    ARCHIVO,
    HOJA
)

X = datos["X"]
Y = datos["Y"]


# ======================================================================
# 5. AJUSTAR EL MODELO OLS
# ======================================================================
#
# Modelo:
#
# Y = β0 + β1 X + ε
#
# β0 = intercepto
# β1 = pendiente
# ε  = error aleatorio
# ======================================================================

X_model = sm.add_constant(X)

modelo = sm.OLS(
    Y,
    X_model
).fit()


# ======================================================================
# 6. COEFICIENTES
# ======================================================================

intercepto = modelo.params.iloc[0]
pendiente = modelo.params.iloc[1]

EE_intercepto = modelo.bse.iloc[0]
EE_pendiente = modelo.bse.iloc[1]

t_intercepto = modelo.tvalues.iloc[0]
t_pendiente = modelo.tvalues.iloc[1]

p_intercepto = modelo.pvalues.iloc[0]
p_pendiente = modelo.pvalues.iloc[1]

IC_coef = modelo.conf_int(
    alpha=ALPHA
)


# ======================================================================
# 7. FIV — FACTOR DE INFLACIÓN DE LA VARIANZA
# ======================================================================
#
# En una regresión lineal SIMPLE sólo existe un predictor.
#
# Por definición:
#
# VIF = 1 / (1 - Rj²)
#
# Como X no puede explicarse mediante otros predictores,
# Rj² = 0 y:
#
# VIF = 1.00
#
# Por eso Minitab muestra FIV = 1.00 para X.
# ======================================================================

FIV_X = 1.0


# ======================================================================
# 8. RESUMEN DEL MODELO
# ======================================================================

S = np.sqrt(
    modelo.ssr /
    modelo.df_resid
)

R2 = modelo.rsquared

R2_ajustado = (
    modelo.rsquared_adj
)

# PRESS mediante residuos eliminados
influencia = modelo.get_influence()

residuos_PRESS = (
    influencia.resid_press
)

PRESS = np.sum(
    residuos_PRESS ** 2
)

TSS = np.sum(
    (Y - Y.mean()) ** 2
)

R2_predicho = (
    1 - PRESS / TSS
    if TSS > 0
    else np.nan
)


# ======================================================================
# 9. ANOVA
# ======================================================================

tabla_anova = calcular_anova(
    modelo
)

p_anova = (
    tabla_anova
    .loc[
        tabla_anova["Fuente"] == "Regresión",
        "p"
    ]
    .iloc[0]
)


# ======================================================================
# 10. LACK-OF-FIT
# ======================================================================

lof = calcular_lack_of_fit(
    datos,
    modelo
)


# ======================================================================
# 11. DIAGNÓSTICOS
# ======================================================================

diagnosticos, umbral_resid, umbral_leverage, umbral_cook = (
    obtener_diagnosticos(
        modelo,
        datos
    )
)


# ======================================================================
# 12. SUPUESTOS
# ======================================================================

# --------------------------------------------------------------
# Normalidad — Anderson-Darling
# --------------------------------------------------------------

AD, p_AD = normal_ad(
    modelo.resid
)


# --------------------------------------------------------------
# Independencia — Durbin-Watson
# --------------------------------------------------------------

DW = durbin_watson(
    modelo.resid
)


# --------------------------------------------------------------
# Homocedasticidad — Breusch-Pagan
# --------------------------------------------------------------

BP_LM, BP_p_LM, BP_F, BP_p_F = (
    het_breuschpagan(
        modelo.resid,
        modelo.model.exog
    )
)


# ======================================================================
# 13. ENCABEZADO
# ======================================================================

print("\n" + "=" * 90)
print("REGRESIÓN LINEAL SIMPLE — MÍNIMOS CUADRADOS ORDINARIOS (OLS)")
print("APLICACIÓN A VALIDACIÓN DE MÉTODOS QUÍMICOS CUANTITATIVOS")
print("=" * 90)

print(f"\nArchivo : {ARCHIVO}")
print(f"Hoja   : {hoja_real}")

print("\nVariables detectadas:")
print(f"X = {nombre_x}")
print(f"Y = {nombre_y}")

print(f"\nN observaciones = {len(datos)}")
print(f"N niveles de X  = {datos['X'].nunique()}")


# ======================================================================
# 14. ECUACIÓN DE REGRESIÓN
# ======================================================================

print("\n" + "-" * 90)
print("1. ECUACIÓN DE REGRESIÓN")
print("-" * 90)

signo = "+" if pendiente >= 0 else "-"

print(
    f"\n{nombre_y} = "
    f"{intercepto:.6f} "
    f"{signo} "
    f"{abs(pendiente):.6f} × {nombre_x}"
)


# ======================================================================
# 15. COEFICIENTES
# ======================================================================

print("\n" + "-" * 90)
print("2. COEFICIENTES DEL MODELO")
print("-" * 90)

tabla_coef = pd.DataFrame({

    "Término": [
        "Constante",
        nombre_x
    ],

    "Coef": [
        intercepto,
        pendiente
    ],

    "EE del coef.": [
        EE_intercepto,
        EE_pendiente
    ],

    "Valor T": [
        t_intercepto,
        t_pendiente
    ],

    "Valor p": [
        p_intercepto,
        p_pendiente
    ],

    "FIV": [
        np.nan,
        FIV_X
    ],

    "IC95 inferior": [
        IC_coef.iloc[0, 0],
        IC_coef.iloc[1, 0]
    ],

    "IC95 superior": [
        IC_coef.iloc[0, 1],
        IC_coef.iloc[1, 1]
    ]
})

print(
    tabla_coef.to_string(
        index=False,
        float_format=lambda x: f"{x:.6g}"
    )
)

print(
    "\nInterpretación didáctica:"
)

print(
    f"• Pendiente: p = {fmt_p(p_pendiente)}. "
    "Evalúa si existe evidencia estadística de una relación lineal."
)

print(
    f"• Constante: p = {fmt_p(p_intercepto)}. "
    "Evalúa si el intercepto difiere estadísticamente de cero."
)

print(
    f"• FIV de X = {FIV_X:.2f}. "
    "En regresión simple es 1.00 porque no existen otros predictores "
    "con los cuales X pueda presentar colinealidad."
)


# ======================================================================
# 16. RESUMEN DEL MODELO
# ======================================================================

print("\n" + "-" * 90)
print("3. RESUMEN DEL MODELO")
print("-" * 90)

print(f"\nS                    = {S:.6f}")
print(f"R-cuadrado           = {R2 * 100:.6f} %")
print(f"R-cuadrado ajustado  = {R2_ajustado * 100:.6f} %")
print(f"R-cuadrado predicho  = {R2_predicho * 100:.6f} %")
print(f"PRESS                = {PRESS:.6f}")

print(
    "\nNota: no existe un umbral universal de R² para declarar "
    "un método lineal. Debe interpretarse junto con Lack-of-Fit, "
    "residuales y capacidad predictiva."
)


# ======================================================================
# 17. ANÁLISIS DE VARIANZA + LACK-OF-FIT
# ======================================================================

print("\n" + "=" * 90)
print("4. ANÁLISIS DE VARIANZA DE LA REGRESIÓN")
print("=" * 90)

# ----------------------------------------------------------------------
# Tabla ANOVA de regresión
#
# Esta es la tabla ANOVA que debe mostrarse explícitamente.
# Separa:
#   Regresión
#   X
#   Error
#   Total
#
# Para una regresión lineal simple:
#
#   GL regresión = 1
#   GL error     = n - 2
#   GL total     = n - 1
#
# El valor F y su p-valor prueban:
#
#   H0: β1 = 0
#   H1: β1 ≠ 0
# ----------------------------------------------------------------------

tabla_anova_presentacion = tabla_anova.copy()

tabla_anova_presentacion["SC"] = tabla_anova_presentacion["SC"].round(6)
tabla_anova_presentacion["MC"] = tabla_anova_presentacion["MC"].round(6)
tabla_anova_presentacion["F"] = tabla_anova_presentacion["F"].round(6)
tabla_anova_presentacion["p"] = tabla_anova_presentacion["p"].round(6)

print("\nANOVA de regresión:")
print(
    tabla_anova_presentacion.to_string(
        index=False,
        float_format=lambda x: f"{x:.6f}"
    )
)

# En Jupyter, además de imprimir, mostrar como tabla visual.
if display is not None:
    display(tabla_anova_presentacion)

print("\nHipótesis del ANOVA de regresión:")
print("H₀: β₁ = 0  → no existe relación lineal significativa.")
print("H₁: β₁ ≠ 0  → existe evidencia de relación lineal significativa.")

f_anova_regresion = tabla_anova.loc[
    tabla_anova["Fuente"] == "Regresión", "F"
].iloc[0]

print(
    f"\nEstadístico F = {f_anova_regresion:.6f}"
)

print(
    f"Valor p = {fmt_p(p_anova)}"
)

if p_anova < ALPHA:
    print(
        "Conclusión: SE RECHAZA H₀. "
        "Existe evidencia estadística de relación lineal."
    )
else:
    print(
        "Conclusión: NO SE RECHAZA H₀. "
        "No existe evidencia estadística suficiente de relación lineal."
    )

print(
    "\nIMPORTANTE: un ANOVA de regresión significativo NO demuestra "
    "por sí solo que el modelo lineal sea adecuado. Deben revisarse "
    "Lack-of-Fit y los supuestos mediante los residuales."
)


# ----------------------------------------------------------------------
# Lack-of-Fit dentro del análisis de varianza
# ----------------------------------------------------------------------

print("\n" + "=" * 90)
print("4.1 LACK-OF-FIT — PRUEBA DE FALTA DE AJUSTE")
print("=" * 90)

# Inicializamos la variable para que la exportación posterior
# no produzca errores cuando no existan réplicas.
tabla_error = None

if lof["hay_replicas"] and lof["df_LOF"] > 0 and lof["df_PE"] > 0:

    # --------------------------------------------------------------
    # Tabla de descomposición del error
    # --------------------------------------------------------------

    tabla_error = pd.DataFrame({

        "Fuente": [
            "Error",
            "  Falta de ajuste",
            "  Error puro"
        ],

        "GL": [
            len(datos) - 2,
            lof["df_LOF"],
            lof["df_PE"]
        ],

        "SC": [
            lof["SSE"],
            lof["SS_LOF"],
            lof["SS_PE"]
        ],

        "MC": [
            lof["SSE"] / (len(datos) - 2),
            lof["MS_LOF"],
            lof["MS_PE"]
        ],

        "F": [
            np.nan,
            lof["F_LOF"],
            np.nan
        ],

        "p": [
            np.nan,
            lof["p_LOF"],
            np.nan
        ]
    })

    tabla_error_presentacion = tabla_error.copy()

    for columna in ["SC", "MC", "F", "p"]:
        tabla_error_presentacion[columna] = (
            tabla_error_presentacion[columna].round(6)
        )

    print("\nDescomposición del error:")
    print(
        tabla_error_presentacion.to_string(
            index=False,
            float_format=lambda x: f"{x:.6f}"
        )
    )

    if display is not None:
        display(tabla_error_presentacion)

    print("\nDescomposición:")
    print("SSE = SS_LOF + SS_PE")

    print(
        f"{lof['SSE']:.6f} = "
        f"{lof['SS_LOF']:.6f} + "
        f"{lof['SS_PE']:.6f}"
    )

    print(
        "\nPrueba de Lack-of-Fit:"
    )

    print(
        f"F = {lof['F_LOF']:.6f}"
    )

    print(
        f"p = {fmt_p(lof['p_LOF'])}"
    )

    print(
        "\nHipótesis:"
    )

    print(
        "H₀: no existe falta de ajuste significativa."
    )

    print(
        "H₁: existe falta de ajuste significativa."
    )

    if lof["p_LOF"] >= ALPHA:

        print(
            "\nConclusión: NO SE RECHAZA H₀. "
            "No se evidencia falta de ajuste estadísticamente significativa."
        )

        print(
            "Interpretación: los datos son compatibles con la forma "
            "lineal propuesta, dentro de la variabilidad experimental "
            "estimada mediante el error puro."
        )

    else:

        print(
            "\nConclusión: SE RECHAZA H₀. "
            "Existe evidencia estadística de falta de ajuste."
        )

        print(
            "Interpretación: la variabilidad respecto a la recta "
            "es mayor que la esperada por el error puro. "
            "Debe investigarse si existe curvatura, cambio de respuesta "
            "con la concentración u otra estructura no explicada por "
            "el modelo lineal."
        )

    print(
        "\nNOTA DIDÁCTICA: el Lack-of-Fit requiere réplicas en niveles "
        "de X para poder estimar el error puro. Sin réplicas, la prueba "
        "no puede separarse estadísticamente en falta de ajuste y error puro."
    )

else:

    print(
        "\nLack-of-Fit NO EVALUABLE."
    )

    print(
        "No existen réplicas suficientes en los niveles de X para "
        "estimar el error puro."
    )

    print(
        "Por tanto, no debe interpretarse la ausencia de esta prueba "
        "como evidencia de buen ajuste."
    )


# ----------------------------------------------------------------------
# Resumen conjunto: ANOVA + Lack-of-Fit
# ----------------------------------------------------------------------

print("\n" + "-" * 90)
print("RESUMEN DE LAS PRUEBAS DE AJUSTE")
print("-" * 90)

resumen_ajuste = pd.DataFrame({
    "Prueba": [
        "ANOVA de regresión",
        "Lack-of-Fit"
    ],

    "Estadístico": [
        tabla_anova.loc[
            tabla_anova["Fuente"] == "Regresión",
            "F"
        ].iloc[0],
        lof["F_LOF"] if np.isfinite(lof["F_LOF"]) else np.nan
    ],

    "p-valor": [
        p_anova,
        lof["p_LOF"] if np.isfinite(lof["p_LOF"]) else np.nan
    ],

    "Criterio": [
        "p < 0.05 → relación lineal significativa",
        "p ≥ 0.05 → no se evidencia falta de ajuste"
    ]
})

if display is not None:
    display(resumen_ajuste)
else:
    print(
        resumen_ajuste.to_string(
            index=False,
            float_format=lambda x: f"{x:.6f}"
        )
    )


# 18. AJUSTES Y DIAGNÓSTICO DE OBSERVACIONES POCO COMUNES
# ======================================================================

print("\n" + "-" * 90)
print("5. AJUSTES Y DIAGNÓSTICO DE OBSERVACIONES POCO COMUNES")
print("-" * 90)

tabla_obs = diagnosticos[[
    "Obs",
    "X",
    "Y",
    "Ajuste",
    "EE ajuste",
    "Resid",
    "Resid est",
    "Leverage",
    "Cook",
    "Marca R",
    "Marca leverage",
    "Marca Cook"
]]

print(
    tabla_obs.to_string(
        index=False,
        float_format=lambda x: f"{x:.6g}"
    )
)

print("\nCriterios diagnósticos utilizados:")

print(
    f"• R: |Residual estandarizado| > {umbral_resid:.1f}"
)

print(
    f"• L: Leverage > 2p/n = {umbral_leverage:.6f}"
)

print(
    f"• C: Cook > 4/n = {umbral_cook:.6f}"
)

print(
    "\nEstos son umbrales de señalización para investigar observaciones; "
    "NO significan que una observación deba eliminarse automáticamente."
)

observaciones_alerta = diagnosticos[
    (
        (diagnosticos["Marca R"] != "") |
        (diagnosticos["Marca leverage"] != "") |
        (diagnosticos["Marca Cook"] != "")
    )
]

if len(observaciones_alerta) > 0:

    print("\n⚠ OBSERVACIONES PARA REVISIÓN:")

    print(
        observaciones_alerta[
            [
                "Obs",
                "X",
                "Y",
                "Resid est",
                "Leverage",
                "Cook",
                "Marca R",
                "Marca leverage",
                "Marca Cook"
            ]
        ].to_string(
            index=False,
            float_format=lambda x: f"{x:.6g}"
        )
    )

else:

    print(
        "\n✓ No se identificaron observaciones que superen "
        "los umbrales diagnósticos establecidos."
    )


# ======================================================================
# 19. SUPUESTO 1 — NORMALIDAD
# ======================================================================

print("\n" + "-" * 90)
print("6. SUPUESTO OLS #1 — NORMALIDAD DE LOS RESIDUALES")
print("-" * 90)

print("Prueba: Anderson-Darling")
print(f"A² = {AD:.6f}")
print(f"p  = {fmt_p(p_AD)}")

if p_AD >= ALPHA:

    print(
        "Conclusión: NO se rechaza H₀. "
        "No se evidencia desviación estadísticamente significativa "
        "respecto a la normalidad."
    )

else:

    print(
        "Conclusión: SE RECHAZA H₀. "
        "Se evidencia desviación estadísticamente significativa "
        "respecto a la normalidad."
    )

print(
    "\nLa prueba se interpreta junto con el Q-Q Plot y el histograma."
)


# ======================================================================
# 20. SUPUESTO 2 — INDEPENDENCIA
# ======================================================================

print("\n" + "-" * 90)
print("7. SUPUESTO OLS #2 — INDEPENDENCIA DE LOS RESIDUALES")
print("-" * 90)

print(
    f"Durbin-Watson = {DW:.6f}"
)

print(
    "Referencia orientativa: valores cercanos a 2 son compatibles "
    "con ausencia de autocorrelación de primer orden."
)

print(
    "\nLa gráfica Residual vs Orden es fundamental porque permite "
    "visualizar tendencias, ciclos o agrupamientos."
)


# ======================================================================
# 21. SUPUESTO 3 — HOMOCEDASTICIDAD
# ======================================================================

print("\n" + "-" * 90)
print("8. SUPUESTO OLS #3 — HOMOCEDASTICIDAD")
print("-" * 90)

print("Diagnóstico complementario: Breusch-Pagan")

print(f"LM = {BP_LM:.6f}")
print(f"p (LM) = {fmt_p(BP_p_LM)}")

print(f"F = {BP_F:.6f}")
print(f"p (F) = {fmt_p(BP_p_F)}")

if BP_p_F >= ALPHA:

    print(
        "Conclusión: no se evidencia heterocedasticidad "
        "estadísticamente significativa."
    )

else:

    print(
        "Conclusión: se evidencia heterocedasticidad "
        "estadísticamente significativa."
    )

print(
    "\nLa prueba debe complementarse con Residual vs Ajustado. "
    "Un patrón de embudo o dispersión creciente es una señal "
    "visual de varianza no constante."
)


# ======================================================================
# 22. MATRIZ FINAL DE EVALUACIÓN
# ======================================================================

print("\n" + "=" * 90)
print("9. MATRIZ FINAL DE EVALUACIÓN")
print("=" * 90)

filas = []

filas.append([
    "Pendiente significativa",
    p_pendiente,
    "< 0.05",
    "CUMPLE" if p_pendiente < ALPHA else "REVISAR"
])

if np.isfinite(lof["p_LOF"]):

    filas.append([
        "Falta de ajuste",
        lof["p_LOF"],
        "≥ 0.05",
        "CUMPLE" if lof["p_LOF"] >= ALPHA else "REVISAR"
    ])

else:

    filas.append([
        "Falta de ajuste",
        np.nan,
        "Requiere réplicas",
        "NO EVALUABLE"
    ])

filas.append([
    "Normalidad AD",
    p_AD,
    "≥ 0.05",
    "CUMPLE" if p_AD >= ALPHA else "REVISAR"
])

filas.append([
    "Homocedasticidad BP",
    BP_p_F,
    "≥ 0.05",
    "CUMPLE" if BP_p_F >= ALPHA else "REVISAR"
])

filas.append([
    "Independencia DW",
    DW,
    "≈ 2",
    (
        "CUMPLE"
        if 1.5 <= DW <= 2.5
        else "REVISAR"
    )
])

tabla_final = pd.DataFrame(
    filas,
    columns=[
        "Criterio",
        "Resultado",
        "Referencia",
        "Evaluación"
    ]
)

print(
    tabla_final.to_string(
        index=False,
        float_format=lambda x: f"{x:.6g}"
    )
)


# ======================================================================
# 23. CONCLUSIÓN AUTOMÁTICA
# ======================================================================

print("\n" + "=" * 90)
print("10. CONCLUSIÓN ESTADÍSTICA")
print("=" * 90)

print(
    f"\nLa pendiente presenta p = {fmt_p(p_pendiente)} "
    "y el modelo presenta "
    f"R² = {R2 * 100:.6f} %."
)

if np.isfinite(lof["p_LOF"]):

    if lof["p_LOF"] >= ALPHA:

        print(
            f"Lack-of-Fit: p = {fmt_p(lof['p_LOF'])}. "
            "No se evidencia falta de ajuste significativa."
        )

    else:

        print(
            f"Lack-of-Fit: p = {fmt_p(lof['p_LOF'])}. "
            "Se evidencia falta de ajuste significativa."
        )

print(
    "\nLa decisión final NO debe basarse exclusivamente en R² "
    "o en el ANOVA de regresión."
)

problemas = []

if p_AD < ALPHA:
    problemas.append("normalidad")

if BP_p_F < ALPHA:
    problemas.append("homocedasticidad")

if DW < 1.5 or DW > 2.5:
    problemas.append("independencia/autocorrelación")

if np.isfinite(lof["p_LOF"]) and lof["p_LOF"] < ALPHA:
    problemas.append("falta de ajuste")

if problemas:

    print(
        "\n⚠ DIAGNÓSTICO: se detectaron señales que requieren "
        "revisión en: "
        + ", ".join(problemas)
        + "."
    )

    print(
        "\nNo se recomienda declarar automáticamente la adecuación "
        "completa del modelo OLS."
    )

    print(
        "Investigar primero la causa experimental/analítica y, "
        "si corresponde, evaluar un modelo alternativo."
    )

else:

    print(
        "\n✓ No se detectaron señales estadísticas importantes "
        "en los criterios y supuestos evaluados."
    )


# ======================================================================
# 24. GRÁFICA 1 — CURVA DE CALIBRACIÓN PROFESIONAL
# ======================================================================
#
# Incluye:
#   • datos experimentales
#   • promedio de cada nivel
#   • recta OLS
#   • IC 95 % de la media
#   • intervalo de predicción 95 %
#   • ecuación
#   • R²
# ======================================================================

x_grid = np.linspace(
    X.min(),
    X.max(),
    500
)

pred_grid = modelo.get_prediction(
    sm.add_constant(x_grid)
)

pred_frame = pred_grid.summary_frame(
    alpha=ALPHA
)

fig = plt.figure(
    figsize=(13, 7.5),
    constrained_layout=True
)

gs = fig.add_gridspec(
    1,
    2,
    width_ratios=[3.8, 1.35]
)

ax = fig.add_subplot(gs[0, 0])
info = fig.add_subplot(gs[0, 1])

# --------------------------------------------------------------
# Intervalo de predicción
# --------------------------------------------------------------

ax.fill_between(
    x_grid,
    pred_frame["obs_ci_lower"],
    pred_frame["obs_ci_upper"],
    color="#90CAF9",
    alpha=0.14,
    label="Intervalo de predicción 95 %"
)

# --------------------------------------------------------------
# Intervalo de confianza de la media
# --------------------------------------------------------------

ax.fill_between(
    x_grid,
    pred_frame["mean_ci_lower"],
    pred_frame["mean_ci_upper"],
    color="#1976D2",
    alpha=0.18,
    label="IC 95 % de la media"
)

# --------------------------------------------------------------
# Recta
# --------------------------------------------------------------

ax.plot(
    x_grid,
    pred_frame["mean"],
    color="#D32F2F",
    linewidth=2.8,
    label="Recta OLS",
    zorder=4
)

# --------------------------------------------------------------
# Datos
# --------------------------------------------------------------

ax.scatter(
    X,
    Y,
    s=54,
    color="#1565C0",
    edgecolor="white",
    linewidth=0.9,
    alpha=0.92,
    label="Datos experimentales",
    zorder=5
)

# --------------------------------------------------------------
# Promedios por nivel
# --------------------------------------------------------------

medias_nivel = (
    datos
    .groupby("X")["Y"]
    .mean()
    .reset_index()
)

ax.scatter(
    medias_nivel["X"],
    medias_nivel["Y"],
    s=115,
    marker="D",
    color="#FF8F00",
    edgecolor="white",
    linewidth=1.0,
    label="Media por nivel",
    zorder=6
)

# --------------------------------------------------------------
# Etiqueta de ecuación
# --------------------------------------------------------------

ecuacion = (
    f"Y = {intercepto:.4f} "
    f"{'+' if pendiente >= 0 else '-'} "
    f"{abs(pendiente):.4f}X"
)

ax.text(
    0.035,
    0.965,
    ecuacion,
    transform=ax.transAxes,
    va="top",
    fontsize=11.5,
    fontweight="bold",
    bbox=dict(
        boxstyle="round,pad=0.55",
        facecolor="white",
        edgecolor="#CFD8DC",
        alpha=0.94
    )
)

# --------------------------------------------------------------
# Apariencia
# --------------------------------------------------------------

ax.set_title(
    "Curva de calibración — Regresión OLS",
    fontsize=17,
    fontweight="bold",
    pad=12
)

ax.set_xlabel(
    nombre_x,
    fontsize=11.5
)

ax.set_ylabel(
    nombre_y,
    fontsize=11.5
)

ax.grid(
    True,
    alpha=0.18,
    linewidth=0.8
)

ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

ax.legend(
    loc="best",
    frameon=True,
    fancybox=True
)

# --------------------------------------------------------------
# Panel estadístico
# --------------------------------------------------------------

info.axis("off")

info.text(
    0.02,
    0.96,
    "RESULTADO ESTADÍSTICO",
    fontsize=14,
    fontweight="bold",
    va="top"
)

texto_info = (
    f"N = {len(datos)}\n\n"
    f"Pendiente = {pendiente:.6f}\n"
    f"p = {fmt_p(p_pendiente)}\n\n"
    f"R² = {R2 * 100:.6f} %\n"
    f"R² ajustado = {R2_ajustado * 100:.6f} %\n"
    f"R² predicho = {R2_predicho * 100:.6f} %\n\n"
    f"S = {S:.6f}\n\n"
    f"AD p = {fmt_p(p_AD)}\n"
    f"DW = {DW:.4f}\n"
    f"BP p = {fmt_p(BP_p_F)}"
)

info.text(
    0.02,
    0.86,
    texto_info,
    fontsize=10.5,
    va="top",
    linespacing=1.25
)

fig.suptitle(
    "Validación de linealidad — modelo lineal simple",
    fontsize=20,
    fontweight="bold"
)

plt.show()


# ======================================================================
# 25. GRÁFICA 2 — PANEL DE DIAGNÓSTICO DE RESIDUALES
# ======================================================================
#
# Cuatro gráficas en una sola figura:
#
#   1. Q-Q Plot
#   2. Residuales vs ajustados
#   3. Residuales vs X
#   4. Residuales vs orden
#
# Esto permite evaluar visualmente los tres supuestos OLS.
# ======================================================================

fig, axes = plt.subplots(
    2,
    2,
    figsize=(14, 9)
)

# ----------------------------------------------------------------------
# 25.1 Q-Q Plot
# ----------------------------------------------------------------------

ax = axes[0, 0]

(
    theoretical,
    ordered_residuals
), (slope_qq, intercept_qq, r_qq) = stats.probplot(
    diagnosticos["Resid"],
    dist="norm"
)

ax.scatter(
    theoretical,
    ordered_residuals,
    s=45,
    color="#1565C0",
    edgecolor="white",
    linewidth=0.7
)

line_x = np.array([
    theoretical.min(),
    theoretical.max()
])

ax.plot(
    line_x,
    intercept_qq + slope_qq * line_x,
    color="#D32F2F",
    linewidth=2
)

ax.set_title(
    f"Q-Q Plot — Normalidad\nAD p = {fmt_p(p_AD)}",
    fontweight="bold"
)

ax.set_xlabel("Cuantiles teóricos")
ax.set_ylabel("Residuales ordenados")

ax.grid(alpha=0.18)


# ----------------------------------------------------------------------
# 25.2 Residuales vs ajustados
# ----------------------------------------------------------------------

ax = axes[0, 1]

ax.scatter(
    diagnosticos["Ajuste"],
    diagnosticos["Resid"],
    s=48,
    color="#1976D2",
    edgecolor="white",
    linewidth=0.7
)

ax.axhline(
    0,
    color="#D32F2F",
    linestyle="--",
    linewidth=1.7
)

ax.set_title(
    f"Residuales vs ajustados\nBreusch-Pagan p = {fmt_p(BP_p_F)}",
    fontweight="bold"
)

ax.set_xlabel("Valores ajustados")
ax.set_ylabel("Residual")

ax.grid(alpha=0.18)


# ----------------------------------------------------------------------
# 25.3 Residuales vs X
# ----------------------------------------------------------------------

ax = axes[1, 0]

ax.scatter(
    diagnosticos["X"],
    diagnosticos["Resid"],
    s=48,
    color="#00796B",
    edgecolor="white",
    linewidth=0.7
)

ax.axhline(
    0,
    color="#D32F2F",
    linestyle="--",
    linewidth=1.7
)

# Media residual por nivel
media_residual = (
    diagnosticos
    .groupby("X")["Resid"]
    .mean()
    .reset_index()
)

ax.plot(
    media_residual["X"],
    media_residual["Resid"],
    color="#FF8F00",
    marker="D",
    linewidth=1.8,
    markersize=6,
    label="Media residual por nivel"
)

ax.set_title(
    "Residuales vs X",
    fontweight="bold"
)

ax.set_xlabel(nombre_x)
ax.set_ylabel("Residual")

ax.grid(alpha=0.18)
ax.legend()


# ----------------------------------------------------------------------
# 25.4 Residuales vs orden
# ----------------------------------------------------------------------

ax = axes[1, 1]

ax.plot(
    diagnosticos["Obs"],
    diagnosticos["Resid"],
    color="#6A1B9A",
    linewidth=1.5,
    alpha=0.75
)

ax.scatter(
    diagnosticos["Obs"],
    diagnosticos["Resid"],
    s=45,
    color="#6A1B9A",
    edgecolor="white",
    linewidth=0.7,
    zorder=3
)

ax.axhline(
    0,
    color="#D32F2F",
    linestyle="--",
    linewidth=1.7
)

ax.set_title(
    f"Residuales vs orden\nDurbin-Watson = {DW:.4f}",
    fontweight="bold"
)

ax.set_xlabel("Orden de observación")
ax.set_ylabel("Residual")

ax.grid(alpha=0.18)

fig.suptitle(
    "Diagnóstico gráfico de los supuestos del modelo OLS",
    fontsize=18,
    fontweight="bold"
)

fig.tight_layout(
    rect=[0, 0, 1, 0.95]
)

plt.show()


# ======================================================================
# 26. GRÁFICA 3 — HISTOGRAMA DE RESIDUALES
# ======================================================================

fig, ax = plt.subplots(
    figsize=(10, 5.8)
)

res = diagnosticos["Resid"].values

ax.hist(
    res,
    bins="auto",
    density=True,
    color="#64B5F6",
    edgecolor="white",
    linewidth=1.0,
    alpha=0.82,
    label="Residuales"
)

media_res = np.mean(res)
sd_res = np.std(
    res,
    ddof=1
)

x_normal = np.linspace(
    res.min(),
    res.max(),
    500
)

if sd_res > 0:

    ax.plot(
        x_normal,
        stats.norm.pdf(
            x_normal,
            media_res,
            sd_res
        ),
        color="#D32F2F",
        linewidth=2.5,
        label="Normal ajustada"
    )

ax.axvline(
    0,
    color="#263238",
    linestyle="--",
    linewidth=1.4
)

ax.set_title(
    f"Histograma de residuales — Anderson-Darling p = {fmt_p(p_AD)}",
    fontsize=16,
    fontweight="bold"
)

ax.set_xlabel("Residual")
ax.set_ylabel("Densidad")

ax.grid(
    axis="y",
    alpha=0.18
)

ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

ax.legend()

plt.tight_layout()
plt.show()


# ======================================================================
# 27. GRÁFICA 4 — OBSERVACIONES POCO COMUNES
# ======================================================================
#
# Dos diagnósticos especialmente útiles:
#
#   • Residual estandarizado
#   • Distancia de Cook
#
# Las líneas son umbrales de señalización.
# ======================================================================

fig, axes = plt.subplots(
    1,
    2,
    figsize=(14, 5.7)
)

# ----------------------------------------------------------------------
# Residuales estandarizados
# ----------------------------------------------------------------------

ax = axes[0]

ax.scatter(
    diagnosticos["Obs"],
    diagnosticos["Resid est"],
    s=52,
    color="#1565C0",
    edgecolor="white",
    linewidth=0.8
)

ax.axhline(
    0,
    color="#263238",
    linestyle="--",
    linewidth=1.2
)

ax.axhline(
    2,
    color="#D32F2F",
    linestyle=":",
    linewidth=1.5
)

ax.axhline(
    -2,
    color="#D32F2F",
    linestyle=":",
    linewidth=1.5
)

ax.axhline(
    3,
    color="#B71C1C",
    linestyle="--",
    linewidth=1.2
)

ax.axhline(
    -3,
    color="#B71C1C",
    linestyle="--",
    linewidth=1.2
)

ax.set_title(
    "Residuales estandarizados",
    fontweight="bold"
)

ax.set_xlabel("Observación")
ax.set_ylabel("Residual estandarizado")

ax.grid(alpha=0.18)


# ----------------------------------------------------------------------
# Cook
# ----------------------------------------------------------------------

ax = axes[1]

ax.stem(
    diagnosticos["Obs"],
    diagnosticos["Cook"],
    linefmt="#1976D2",
    markerfmt="o",
    basefmt=" "
)

ax.axhline(
    umbral_cook,
    color="#D32F2F",
    linestyle="--",
    linewidth=1.5,
    label=f"4/n = {umbral_cook:.4f}"
)

ax.set_title(
    "Distancia de Cook",
    fontweight="bold"
)

ax.set_xlabel("Observación")
ax.set_ylabel("Cook's Distance")

ax.grid(alpha=0.18)
ax.legend()

fig.suptitle(
    "Diagnóstico de observaciones potencialmente influyentes",
    fontsize=17,
    fontweight="bold"
)

fig.tight_layout(
    rect=[0, 0, 1, 0.94]
)

plt.show()


# ======================================================================
# 28. EXPORTAR RESULTADOS A EXCEL
# ======================================================================

with pd.ExcelWriter(
    ARCHIVO_SALIDA,
    engine="openpyxl"
) as writer:

    # Datos originales
    datos.to_excel(
        writer,
        sheet_name="Datos",
        index=False
    )

    # Descriptiva por nivel
    tabla_descriptiva = (
        datos
        .groupby("X")["Y"]
        .agg(
            N="size",
            Media="mean",
            Desv_est="std",
            Mínimo="min",
            Máximo="max"
        )
        .reset_index()
    )

    tabla_descriptiva.to_excel(
        writer,
        sheet_name="Descriptiva",
        index=False
    )

    # Coeficientes
    tabla_coef.to_excel(
        writer,
        sheet_name="Coeficientes",
        index=False
    )

    # Resumen del modelo
    resumen_modelo = pd.DataFrame({
        "Estadístico": [
            "S",
            "R-cuadrado",
            "R-cuadrado ajustado",
            "R-cuadrado predicho",
            "PRESS"
        ],
        "Valor": [
            S,
            R2,
            R2_ajustado,
            R2_predicho,
            PRESS
        ]
    })

    resumen_modelo.to_excel(
        writer,
        sheet_name="Resumen_modelo",
        index=False
    )

    # ANOVA
    tabla_anova.to_excel(
        writer,
        sheet_name="ANOVA",
        index=False
    )

    # Lack-of-Fit
    if tabla_error is not None:

        tabla_error.to_excel(
            writer,
            sheet_name="Lack_of_Fit",
            index=False
        )

    # Diagnóstico de observaciones
    diagnosticos.to_excel(
        writer,
        sheet_name="Diagnostico",
        index=False
    )

    # Diagnósticos de supuestos
    tabla_supuestos = pd.DataFrame({

        "Prueba": [
            "Anderson-Darling",
            "Durbin-Watson",
            "Breusch-Pagan LM",
            "Breusch-Pagan F"
        ],

        "Estadístico": [
            AD,
            DW,
            BP_LM,
            BP_F
        ],

        "p-valor": [
            p_AD,
            np.nan,
            BP_p_LM,
            BP_p_F
        ]
    })

    tabla_supuestos.to_excel(
        writer,
        sheet_name="Supuestos",
        index=False
    )

    # Matriz final
    tabla_final.to_excel(
        writer,
        sheet_name="Matriz_final",
        index=False
    )


# ======================================================================
# 29. FORMATO BÁSICO DEL EXCEL EXPORTADO
# ======================================================================

try:

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule

    wb = load_workbook(
        ARCHIVO_SALIDA
    )

    for ws in wb.worksheets:

        # Congelar encabezado
        ws.freeze_panes = "A2"

        # Encabezados
        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="1565C0"
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # Ajustar ancho
        for columna in ws.columns:

            max_len = 0

            letra = columna[0].column_letter

            for celda in columna:

                if celda.value is not None:

                    max_len = max(
                        max_len,
                        len(str(celda.value))
                    )

            ws.column_dimensions[
                letra
            ].width = min(
                max_len + 2,
                28
            )

    wb.save(
        ARCHIVO_SALIDA
    )

except Exception as e:

    print(
        "\nAdvertencia: no se pudo aplicar "
        f"el formato avanzado al Excel: {e}"
    )


# ======================================================================
# 30. FINAL
# ======================================================================

print("\n" + "=" * 90)
print("ANÁLISIS OLS COMPLETADO")
print("=" * 90)

print(
    f"\nResultados exportados a:"
)

print(
    f"→ {ARCHIVO_SALIDA}"
)

print(
    "\nEl archivo contiene:"
)

print(
    "✓ Datos"
)
print(
    "✓ Estadística descriptiva"
)
print(
    "✓ Coeficientes + p-valor + FIV"
)
print(
    "✓ Resumen del modelo"
)
print(
    "✓ ANOVA de regresión (Regresión, X, Error y Total)"
)
print(
    "✓ Lack-of-Fit + Error puro, cuando existen réplicas"
)
print(
    "✓ Ajustes y diagnóstico de observaciones"
)
print(
    "✓ Supuestos OLS"
)
print(
    "✓ Matriz final de evaluación"
)

print(
    "\nRecordatorio:"
)

print(
    "La detección estadística de una observación poco común "
    "NO implica eliminarla. Primero debe investigarse su causa "
    "experimental, instrumental o documental."
)
